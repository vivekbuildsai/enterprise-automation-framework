from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from framework.enums import Environment
from framework.exceptions import TestDataError

_PROJECT_ROOT = Path(__file__).resolve().parents[2]
_TESTDATA_DIR = _PROJECT_ROOT / "data" / "testdata"


class TestDataLoader:
    """Loads static test data from `data/testdata/` in whichever format the
    module's authors chose to store it — JSON for structured records, CSV
    for tabular bulk data, Excel for data QA hands off already maintaining
    in a spreadsheet. All three return plain Python data (`dict`/`list`),
    never a framework-specific type, so callers build Pydantic models or
    dataclasses from it however suits the module.
    """

    @staticmethod
    def load_json(relative_path: str) -> Any:
        path = _resolve(relative_path)
        return json.loads(path.read_text(encoding="utf-8"))

    @staticmethod
    def load_csv(relative_path: str) -> list[dict[str, str]]:
        path = _resolve(relative_path)
        with path.open(newline="", encoding="utf-8") as fh:
            return list(csv.DictReader(fh))

    @staticmethod
    def load_excel(relative_path: str, *, sheet_name: str | None = None) -> list[dict[str, Any]]:
        path = _resolve(relative_path)
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        if sheet is None:
            raise TestDataError(f"Excel file '{path}' has no active sheet")

        rows = sheet.iter_rows(values_only=True)
        headers = [str(h) for h in next(rows)]
        return [dict(zip(headers, row, strict=False)) for row in rows]

    @staticmethod
    def load_for_environment(module: str, environment: Environment) -> Any:
        """Loads `data/testdata/<module>/<environment>.json` — the
        environment-specific data convention every module follows (e.g.
        different subscriber IDs seeded in DEV vs. QA).
        """
        return TestDataLoader.load_json(f"{module}/{environment.value}.json")


def _resolve(relative_path: str) -> Path:
    path = _TESTDATA_DIR / relative_path
    if not path.exists():
        raise TestDataError(f"No test data file at '{path}' (looked under {_TESTDATA_DIR})")
    return path
