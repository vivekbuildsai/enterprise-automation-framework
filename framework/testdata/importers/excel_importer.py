from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from framework.exceptions import TestDataError


class ExcelImporter:
    """Imports an arbitrary Excel file — see `CsvImporter`'s docstring for
    when to prefer this over `DatasetLoader.load_excel`.
    """

    @staticmethod
    def import_file(path: str | Path, *, sheet_name: str | None = None) -> list[dict[str, Any]]:
        file_path = Path(path)
        if not file_path.exists():
            raise TestDataError(f"No Excel file at '{file_path}'")
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        if sheet is None:
            raise TestDataError(f"Excel file '{file_path}' has no active sheet")

        rows = sheet.iter_rows(values_only=True)
        headers = [str(h) for h in next(rows)]
        return [dict(zip(headers, row, strict=False)) for row in rows]
